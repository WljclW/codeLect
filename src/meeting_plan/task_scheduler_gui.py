"""
任务轮次调度器 - GUI版本
双击运行，通过弹窗交互完成调度并生成Excel文件。
"""

import os
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def get_meeting_days(start_day):
    return {
        start_day: "A",
        start_day + 10: "B",
        start_day + 11: "B",
        start_day + 14: "C",
    }


def schedule_rounds(total_days, num_rounds):
    occupied_days = set()
    schedule = []

    for r in range(num_rounds):
        placed = False
        for start in range(1, total_days + 1):
            meetings = get_meeting_days(start)
            meeting_days = set(meetings.keys())

            if max(meeting_days) > total_days:
                break

            if meeting_days & occupied_days:
                continue

            occupied_days |= meeting_days
            schedule.append({
                "round": r + 1,
                "start_day": start,
                "meetings": meetings,
            })
            placed = True
            break

        if not placed:
            return False, f"无法安排第{r + 1}轮（共需{num_rounds}轮），总天数{total_days}天不够"

    return True, schedule


def export_to_excel(schedule, total_days, num_rounds, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "时间安排表"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    a_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    b_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    c_fill = PatternFill(start_color="4ECB71", end_color="4ECB71", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["天数"]
    for r in range(1, num_rounds + 1):
        headers.append(f"第{r}轮")
    headers.append("备注")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    day_schedule = {}
    for item in schedule:
        for day, phase in item["meetings"].items():
            day_schedule.setdefault(day, []).append((item["round"], phase))

    for day in range(1, total_days + 1):
        row = day + 1
        ws.cell(row=row, column=1, value=f"第{day}天").border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        remarks = []
        for r in range(1, num_rounds + 1):
            col = r + 1
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            entry = None
            if day in day_schedule:
                for rnd, phase in day_schedule[day]:
                    if rnd == r:
                        entry = phase
                        break

            if entry == "A":
                cell.value = "A会议"
                cell.fill = a_fill
                cell.font = Font(bold=True)
                remarks.append(f"第{r}轮A环节会议")
            elif entry == "B":
                cell.value = "B会议"
                cell.fill = b_fill
                cell.font = Font(bold=True)
                remarks.append(f"第{r}轮B环节会议")
            elif entry == "C":
                cell.value = "C会议"
                cell.fill = c_fill
                cell.font = Font(bold=True)
                remarks.append(f"第{r}轮C环节会议")
            else:
                cell.value = ""

        remark_cell = ws.cell(row=row, column=num_rounds + 2)
        remark_cell.value = "；".join(remarks) if remarks else "可并行其他工作"
        remark_cell.border = thin_border

    ws.column_dimensions['A'].width = 10
    for r in range(1, num_rounds + 1):
        col_letter = chr(ord('A') + r) if r < 26 else 'A'
        ws.column_dimensions[col_letter].width = 12
    last_col = chr(ord('A') + num_rounds + 1) if num_rounds + 1 < 26 else 'A'
    ws.column_dimensions[last_col].width = 30

    summary_row = total_days + 3
    ws.cell(row=summary_row, column=1, value="汇总信息").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=1, value=f"总天数：{total_days}天")
    ws.cell(row=summary_row + 2, column=1, value=f"总轮数：{num_rounds}轮")
    ws.cell(row=summary_row + 3, column=1, value="各轮起始日：")
    for i, item in enumerate(schedule):
        ws.cell(row=summary_row + 4 + i, column=1,
                value=f"  第{item['round']}轮：起始第{item['start_day']}天，"
                      f"A=第{item['start_day']}天，"
                      f"B=第{item['start_day']+10}&{item['start_day']+11}天，"
                      f"C=第{item['start_day']+14}天")

    wb.save(filename)


class SchedulerApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("任务轮次调度器")
        self.root.geometry("420x350")
        self.root.resizable(False, False)

        self._build_ui()
        self.root.mainloop()

    def _build_ui(self):
        # 标题
        tk.Label(self.root, text="任务轮次调度器", font=("Microsoft YaHei", 16, "bold")).pack(pady=(20, 5))

        # 规则说明
        rule_text = (
            "规则：每轮含A/B/C三个环节\n"
            "  A环节：轮次第1天开会\n"
            "  B环节：轮次第11、12天开会\n"
            "  C环节：轮次第15天开会\n"
            "约束：不同轮的会议日不能安排在同一天"
        )
        tk.Label(self.root, text=rule_text, justify="left", font=("Microsoft YaHei", 9),
                 fg="#555555").pack(pady=(0, 15))

        # 输入区域
        input_frame = tk.Frame(self.root)
        input_frame.pack(pady=5)

        tk.Label(input_frame, text="总天数：", font=("Microsoft YaHei", 11)).grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.entry_days = tk.Entry(input_frame, width=15, font=("Microsoft YaHei", 11))
        self.entry_days.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(input_frame, text="轮  数：", font=("Microsoft YaHei", 11)).grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.entry_rounds = tk.Entry(input_frame, width=15, font=("Microsoft YaHei", 11))
        self.entry_rounds.grid(row=1, column=1, padx=5, pady=5)

        # 按钮
        tk.Button(self.root, text="开始排程", font=("Microsoft YaHei", 12, "bold"),
                  width=16, height=1, bg="#4472C4", fg="white",
                  command=self._on_run).pack(pady=20)

    def _on_run(self):
        days_str = self.entry_days.get().strip()
        rounds_str = self.entry_rounds.get().strip()

        if not days_str or not rounds_str:
            messagebox.showwarning("输入不完整", "请输入总天数和轮数。")
            return

        try:
            total_days = int(days_str)
            num_rounds = int(rounds_str)
        except ValueError:
            messagebox.showerror("输入错误", "总天数和轮数必须为正整数。")
            return

        if total_days <= 0 or num_rounds <= 0:
            messagebox.showerror("输入错误", "总天数和轮数必须为正整数。")
            return

        success, result = schedule_rounds(total_days, num_rounds)

        if not success:
            messagebox.showinfo("无法实现", result)
            return

        # 成功，让用户选择保存位置
        detail_lines = []
        for item in result:
            s = item['start_day']
            detail_lines.append(
                f"第{item['round']}轮：A=第{s}天, B=第{s+10}&{s+11}天, C=第{s+14}天"
            )
        detail = "\n".join(detail_lines)

        save_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="task_schedule.xlsx"
        )

        if not save_path:
            return

        export_to_excel(result, total_days, num_rounds, save_path)

        msg = f"排程成功！安排如下：\n\n{detail}\n\n文件已保存至：\n{save_path}"
        if messagebox.askyesno("完成", msg + "\n\n是否立即打开文件？"):
            os.startfile(save_path)


if __name__ == "__main__":
    SchedulerApp()
